from os import name
import openpyxl
import datetime

book = openpyxl.load_workbook('data.xlsx')
sheet = book['Sheet1']
#print(sheet.max_row,'|',sheet.max_column)

ls=[]
for i in range(1,2):
    row = ""
    for j in range(1,sheet.max_column):
        row = f'{row}{str(sheet.cell(i,j).value)}\t'
    ls.append(row)
print(ls)

class num:
    name = ""
    week = ""
    enabletime = 0


class detail:
    data = ''#年月日 星期
    startmodule = ''#上班情况 无需打卡、正常、缺卡、
    starttime = ''#上班打卡

    endmodule = ''#上班情况 无需打卡、正常、缺卡、
    endtime = ''#下班打卡

    todaytime = 0 #当日有效工时
    def checktime():
        todaytime = 1

#a1= sheet['A1']
#print(f'{a1.row},{a1.column},{a1.coordinate}')

#print(sheet['A1:B4'])
